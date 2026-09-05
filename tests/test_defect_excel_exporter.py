# -*- coding: utf-8 -*-
"""单元测试 for core.defect_excel_exporter（Excel 导出）。"""
import sys
import os
import tempfile
import shutil

CURRENT_FILE = os.path.abspath(__file__)
PROJECT_ROOT = os.path.dirname(os.path.dirname(CURRENT_FILE))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

import unittest
import openpyxl
from core.defect_excel_exporter import (
    export_defects_xlsx,
    defect_to_standard_row,
    DEFECT_CATEGORY_MAP,
)


class TestDefectToStandardRow(unittest.TestCase):
    """测试 defect_to_standard_row 字段映射。"""

    def test_category_mapping(self):
        """四类缺陷应正确映射到标准分类。"""
        defect = {
            "equip_id": "TMP001",
            "defect_type": "图上有模型无",
            "description": "SVG存在设备",
            "suggestion": "建议补录",
            "sql_draft": "INSERT ...",
        }
        row = defect_to_standard_row(defect, 1, "LINE215", "ST001")
        self.assertEqual(row["序号"], 1)
        self.assertEqual(row["一级分类"], "2 图模一致性校验")
        self.assertEqual(row["二级分类"], "2.1 图上有、模型无校验任务")
        self.assertEqual(row["问题设备id"], "TMP001")
        self.assertEqual(row["所属馈线"], "LINE215")
        self.assertEqual(row["所属厂站"], "ST001")

    def test_model_missing_is_mapped(self):
        defect = {
            "equip_id": "TMP002",
            "defect_type": "模型有图上无",
            "description": "模型有图无",
            "suggestion": "建议补图",
            "sql_draft": "-- 补图",
        }
        row = defect_to_standard_row(defect, 2, "LINE215", "ST001")
        self.assertEqual(row["二级分类"], "2.2 模型有、图上无校验任务")

    def test_unknown_type_uses_default(self):
        defect = {
            "equip_id": "TMP003",
            "defect_type": "未知类型",
            "description": "...",
            "suggestion": "...",
            "sql_draft": "",
        }
        row = defect_to_standard_row(defect, 3, "LINE216", "")
        self.assertEqual(row["一级分类"], "2 图模一致性校验")
        self.assertEqual(row["二级分类"], "2.1 图上有、模型无校验任务")


class TestStandardHeaders(unittest.TestCase):
    """测试表头与标准模板对齐。"""

    def test_10_columns(self):
        """Sheet1 表头为 10 列。"""
        # 从标准模板文件读取验证
        from config.settings import DATASET_STANDARD_OUTPUT_XLSX
        import openpyxl
        if os.path.exists(DATASET_STANDARD_OUTPUT_XLSX):
            wb = openpyxl.load_workbook(DATASET_STANDARD_OUTPUT_XLSX)
            ws = wb["拓扑校验问题清单"]
            headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            wb.close()
            self.assertEqual(len(headers), 10)
        else:
            self.skipTest("标准模板文件不存在")


class TestExportXlsx(unittest.TestCase):
    """测试 export_defects_xlsx 基本功能。"""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def tearDown(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_exports_all_sheets(self):
        from config.settings import DATASET_STANDARD_OUTPUT_XLSX
        if not os.path.exists(DATASET_STANDARD_OUTPUT_XLSX):
            self.skipTest("标准模板文件不存在，跳过集成测试")

        defects = [
            {
                "equip_id": "TMP001",
                "defect_type": "图上有模型无",
                "description": "SVG存在设备但模型缺失",
                "suggestion": "建议补录",
                "sql_draft": "INSERT INTO EQUIP...",
            }
        ]
        out_path = os.path.join(self.tmpdir, "test_report.xlsx")
        result = export_defects_xlsx(
            defects,
            out_path,
            "LINE215",
            template_path=DATASET_STANDARD_OUTPUT_XLSX,
            default_station="ST001",
        )
        self.assertTrue(os.path.exists(result))

        wb = openpyxl.load_workbook(result)
        # Sheet1 表头正确
        ws1 = wb["拓扑校验问题清单"]
        headers = [ws1.cell(1, c).value for c in range(1, 11)]
        expected = ["序号", "一级分类", "二级分类", "问题设备id", "问题设备名称",
                    "所属馈线", "所属厂站", "问题说明", "修正方案", "修正sql"]
        self.assertEqual(headers, expected)
        # 数据行
        row2 = [ws1.cell(2, c).value for c in range(1, 11)]
        self.assertEqual(row2[0], 1)  # 序号 = 1
        self.assertEqual(row2[3], "TMP001")  # 问题设备id

        # Sheet3 联络开关（即使是空数据也要有表头）
        ws3 = wb["联络开关自动识别与可视化梳理任务结果"]
        h3 = [ws3.cell(1, c).value for c in range(1, ws3.max_column + 1)]
        self.assertIn("线路id", h3)
        self.assertIn("联络开关id", h3)

        wb.close()


if __name__ == "__main__":
    unittest.main()
